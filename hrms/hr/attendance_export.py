# Copyright (c) 2026, Frappe Technologies Pvt. Ltd. and Contributors
# License: GNU General Public License v3. See license.txt

"""The attendance grid as a spreadsheet, for the page and for the report.

Both screens hand over the same shape — a row per employee, a cell per day, the marks and
colours of `attendance_marks` — and this module turns that shape into a workbook. Where
the days come from stays with whoever asks: the page exports its live sheet, the report
exports what payroll has been handed, and neither can print the other's numbers.

The frozen pane over the first column is the spreadsheet's answer to the sticky name
column both screens keep: a month is wider than any window.
"""

import json
from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import frappe
from frappe import _
from frappe.desk.utils import provide_binary_file
from frappe.utils import cstr, flt, getdate

from hrms.hr.attendance_marks import get_abbr, get_color, get_day_label

REPORT = "Monthly Attendance Sheet"

# the marks carry the colours of the screen, and openpyxl speaks only in hex
COLOR_NAMES = {"green": "2E7D32", "red": "C62828"}

# a day nobody was meant to work, in the tint the sheet page paints it with
OFF_FILL = "FFD1D1"

HEADER_FILL = "F4F4F5"

NOTE_COLORS = {"over": "2E7D32", "under": "C62828", "leave": "878787"}

NAME_WIDTH = 34

DAY_WIDTH = 6

# the totals of the summarized view, in the order both screens read them
TOTAL_FIELDS = (
	("total_present", "Present Days"),
	("total_leave", "Leave Days"),
	("total_sick", "Sick Days"),
	("total_absent", "Absent Days"),
	("overtime_hours", "Overtime Hours"),
	("shortfall_hours", "Shortfall Hours"),
)

PAID_STATUSES = ("Present", "Work From Home")

# the days nobody was meant to work, tinted rather than counted
NON_WORKING_STATUSES = ("Weekly Off", "Holiday")


@frappe.whitelist()
def download_sheet(company: str, from_date: str, to_date: str) -> None:
	"""The attendance sheet page, as its reader sees it — unapproved days and all."""
	from hrms.hr.page.attendance_sheet.attendance_sheet import build_sheet, validate_period

	from_date, to_date = validate_period(from_date, to_date)
	sheet = build_sheet(company, from_date, to_date)

	if not sheet["employees"]:
		frappe.throw(_("Nothing to export for this period."))

	provide_workbook(get_grid_of_sheet(sheet, company, from_date, to_date), from_date, to_date)


@frappe.whitelist()
def download_report(filters: str) -> None:
	"""The monthly report, with the filters it is being read under."""
	filters = frappe._dict(json.loads(filters) if isinstance(filters, str) else filters or {})

	assert_can_read_report()

	grid = get_grid_of_report(filters)

	if not grid["rows"]:
		frappe.throw(_("Nothing to export for this period."))

	dates = grid["days"]

	provide_workbook(grid, dates[0]["date"], dates[-1]["date"])


def assert_can_read_report() -> None:
	"""The export is the report by another route, so it asks for the same permission."""
	if not frappe.get_doc("Report", REPORT).is_permitted():
		raise frappe.PermissionError(_("Not permitted to read this report"))

	frappe.has_permission("Attendance", throw=True)


# ------------------------------------------------------------------ the grid


def get_grid_of_sheet(sheet: dict, company: str, from_date, to_date) -> dict:
	"""The page's own rows: a cell per day, and the totals it counts from those cells."""
	dates = [getdate(day) for day in sheet["dates"]]
	rows = [
		{
			"name": row["employee_name"] or row["employee"],
			"cells": [get_cell_of_sheet(row["days"].get(cstr(day)) or {}) for day in dates],
			"totals": get_totals_of_sheet(row),
		}
		for row in sheet["employees"]
	]

	return {
		"title": get_title(company, from_date, to_date),
		"days": [{"date": day, "label": get_day_label(day)} for day in dates],
		"rows": rows,
	}


def get_cell_of_sheet(cell: dict) -> dict:
	"""A day of the page: the mark in the colour of its status, the hours or leave under it."""
	status = cell.get("status") or ""

	return {
		"text": get_abbr(status) if status else "",
		"color": get_color(status) if status else None,
		"note": get_note_of_sheet(cell),
		"off": status in NON_WORKING_STATUSES,
	}


def get_note_of_sheet(cell: dict) -> dict | None:
	"""The second line of a day, in the order the page draws it: hours first, leave after."""
	if flt(cell.get("overtime_hours")):
		return {"text": f"+{flt(cell['overtime_hours']):g}", "kind": "over"}

	if flt(cell.get("shortfall_hours")):
		return {"text": f"-{flt(cell['shortfall_hours']):g}", "kind": "under"}

	return {"text": cell["leave_abbr"], "kind": "leave"} if cell.get("leave_abbr") else None


def get_totals_of_sheet(row: dict) -> dict[str, float]:
	"""The page counts its totals from the cells on screen, and so does its export."""
	totals = dict.fromkeys((field for field, _label in TOTAL_FIELDS), 0.0)

	for cell in row["days"].values():
		status = cell.get("status")

		if status in PAID_STATUSES:
			totals["total_present"] += 1
		elif status == "On Leave":
			# a leave nobody pays for is an absence at the employee's own expense
			totals["total_absent" if cell.get("unpaid_leave") else "total_leave"] += 1
		elif status == "Sick Leave":
			totals["total_sick"] += 1
		elif status == "Absent":
			totals["total_absent"] += 1

		totals["overtime_hours"] += flt(cell.get("overtime_hours"))
		totals["shortfall_hours"] += flt(cell.get("shortfall_hours"))

	return totals


def get_grid_of_report(filters: dict) -> dict:
	"""The report's own rows, read twice: once for the days, once for the totals.

	The totals come from the summarized view rather than from the days above them, because
	that is where the report itself gets them — an export that counted its own would be a
	second opinion nobody asked for.
	"""
	from hrms.hr.report.monthly_attendance_sheet.monthly_attendance_sheet import (
		execute,
		get_dates_in_period,
	)

	filters = frappe._dict(filters)
	filters.summarized_view = 0
	filters.unsubmitted_view = 0
	filters.show_chart = 0

	_columns, data, _message, _chart = execute(filters)
	days = [getdate(day) for day in get_dates_in_period(filters)]
	totals = get_totals_of_report(filters)

	rows = [
		{
			"name": get_report_row_name(row),
			"heading": not row.get("employee"),
			"cells": [get_cell_of_report(row, day) for day in days],
			"totals": totals.get(row.get("employee")),
		}
		for row in data
	]

	return {
		"title": get_title(filters.company, days[0], days[-1]),
		"days": [{"date": day, "label": get_day_label(day)} for day in days],
		"rows": rows,
	}


def get_report_row_name(row: dict) -> str:
	"""The name of a report row, with the shift after it when the employee works more than one."""
	if not row.get("employee"):
		# a grouping heading carries the value it groups under and nothing else
		return cstr(next(iter(row.values()), ""))

	name = row.get("employee_name") or row["employee"]

	return f"{name} · {row['shift']}" if row.get("shift") else name


def get_cell_of_report(row: dict, day: date) -> dict:
	"""A day of the report, taken from the marks it already carries."""
	fieldname = day.strftime("%d-%m-%Y")
	mark = (row.get("marks") or {}).get(fieldname) or {}

	return {
		"text": row.get(fieldname) or "",
		"color": mark.get("color"),
		"note": mark.get("note"),
		"off": bool(mark.get("off")),
	}


def get_totals_of_report(filters: dict) -> dict[str, dict]:
	"""The summarized view of the same period, keyed by employee."""
	from hrms.hr.report.monthly_attendance_sheet.monthly_attendance_sheet import execute

	summarized = frappe._dict(filters)
	summarized.summarized_view = 1

	_columns, data, _message, _chart = execute(summarized)

	return {row["employee"]: row for row in data if row.get("employee")}


def get_title(company: str, from_date, to_date) -> str:
	return _("Attendance Sheet: {0}, {1} — {2}").format(
		company,
		frappe.format(from_date, {"fieldtype": "Date"}),
		frappe.format(to_date, {"fieldtype": "Date"}),
	)


# ------------------------------------------------------------------ the workbook


def provide_workbook(grid: dict, from_date, to_date) -> None:
	# the name is kept in plain characters: it travels in a response header
	filename = f"attendance-{getdate(from_date)}-{getdate(to_date)}"

	provide_binary_file(filename, "xlsx", write_grid(grid).getvalue())


def write_grid(grid: dict) -> BytesIO:
	"""The grid as a sheet: days across, employees down, totals below them."""
	workbook = Workbook()
	sheet = workbook.active
	sheet.title = _("Attendance")

	write_title(sheet, grid)
	write_head(sheet, grid)
	write_days(sheet, grid)
	write_totals(sheet, grid)
	set_widths(sheet, grid)

	# everything above the first employee and left of the first day stays in sight
	sheet.freeze_panes = "B3"

	stream = BytesIO()
	workbook.save(stream)
	stream.seek(0)

	return stream


def write_title(sheet, grid: dict) -> None:
	cell = sheet.cell(row=1, column=1, value=grid["title"])
	cell.font = Font(bold=True, size=12)


def write_head(sheet, grid: dict) -> None:
	head = [_("Employee")] + [day["label"] for day in grid["days"]]

	for column, label in enumerate(head, start=1):
		cell = sheet.cell(row=2, column=column, value=label)
		cell.font = Font(bold=True)
		cell.alignment = Alignment(horizontal="left" if column == 1 else "center", wrap_text=True)
		cell.fill = PatternFill("solid", fgColor=HEADER_FILL)


def write_days(sheet, grid: dict) -> None:
	for index, row in enumerate(grid["rows"], start=3):
		name = sheet.cell(row=index, column=1, value=row["name"])
		name.font = Font(bold=bool(row.get("heading")))
		name.alignment = Alignment(vertical="center")

		for column, cell in enumerate(row["cells"], start=2):
			write_day(sheet.cell(row=index, column=column), cell)


def write_day(target, cell: dict) -> None:
	"""One day: the mark, and under it the line the screen puts under it."""
	note = cell.get("note") or {}
	target.value = "\n".join(part for part in (cell["text"], note.get("text")) if part)
	target.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

	if cell.get("color"):
		target.font = Font(bold=True, color=get_argb(cell["color"]))

	if cell.get("off"):
		target.fill = PatternFill("solid", fgColor=OFF_FILL)


def write_totals(sheet, grid: dict) -> None:
	"""The summary block under the days, for the reader who scrolled to the end of the month."""
	rows = [row for row in grid["rows"] if row.get("totals")]

	if not rows:
		return

	start = len(grid["rows"]) + 5
	sheet.cell(row=start - 1, column=1, value=_("Summary")).font = Font(bold=True, size=12)

	head = [_("Employee")] + [_(label) for _field, label in TOTAL_FIELDS]

	for column, label in enumerate(head, start=1):
		cell = sheet.cell(row=start, column=column, value=label)
		cell.font = Font(bold=True)
		cell.alignment = Alignment(horizontal="left" if column == 1 else "center", wrap_text=True)
		cell.fill = PatternFill("solid", fgColor=HEADER_FILL)

	for index, row in enumerate(rows, start=start + 1):
		sheet.cell(row=index, column=1, value=row["name"])

		for column, (field, _label) in enumerate(TOTAL_FIELDS, start=2):
			total = sheet.cell(row=index, column=column, value=flt(row["totals"].get(field)))
			total.alignment = Alignment(horizontal="center")


def set_widths(sheet, grid: dict) -> None:
	"""The name column carries a name, the rest carry a letter or a number.

	The summary block sits under the days and shares their columns, so its headings are
	wrapped rather than widened: the month is what the sheet is read across.
	"""
	sheet.column_dimensions["A"].width = NAME_WIDTH

	for column in range(2, max(len(grid["days"]), len(TOTAL_FIELDS)) + 2):
		sheet.column_dimensions[get_column_letter(column)].width = DAY_WIDTH


def get_argb(color: str) -> str:
	"""A CSS colour of the marks in the notation a spreadsheet understands."""
	if color in COLOR_NAMES:
		return COLOR_NAMES[color]

	return color.lstrip("#").upper()
